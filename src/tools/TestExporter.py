# **************************************************************************************************
# @file TextExporter.py
# @brief Used to create an Excel (.xlsl) file from the test results for them to be printed.
#
# @project   VVToolkit
# @version   1.0
# @date      2024-08-10
# @author    @dabecart
#
# @license
# This project is licensed under the MIT License - see the LICENSE file for details.
# **************************************************************************************************

import os
import shutil
import re
import zipfile
import xml.etree.ElementTree as ET

import openpyxl
import openpyxl.cell
import openpyxl.worksheet
from openpyxl.styles import PatternFill, Border, Alignment, Protection, Font

from DataFields import Item, TestDataFields, TestResult

from typing import List

from PyQt6.QtCore import QFile, QIODevice
from io import BytesIO

# Don't remove this "unused" import, contains the resource images.
import ResourcePacket

def replacePlaceholders(filePath: str, testFields: TestDataFields, items: List[Item]):
    # Load the Excel file from ResourcePacket.
    qfile = QFile(':excel-model')
    if not qfile.open(QIODevice.OpenModeFlag.ReadOnly):
        raise IOError("Failed to open the embedded Excel file")

    excelData = qfile.readAll()
    qfile.close()

    # Duplicate the model into filePath.
    with open(filePath, 'wb') as f:
        f.write(excelData)

    # Convert to a BytesIO object
    excelFileStream = BytesIO(excelData.data())

    modelWorkbook = openpyxl.load_workbook(excelFileStream)
    # Select the worksheet.
    modelSheet = modelWorkbook["VFR"]
    
    # This does not work! And I really need it :(
    # modelSheet.print_area = "A:B"
    # Not even this works!
    # modelSheet.print_area = "A1:B2"

    # As I cannot set the columns as the print area, I'll set a random range inside the excel file.
    # This will create the field inside the workbook.xml file contained in the .xlsx (an .xlsx is 
    # basically a .zip file) and then I'll substitute it with "$A:$B" which is what i want. 

    destinyWorkbook = openpyxl.load_workbook(filePath)
    destinySheet = destinyWorkbook["VFR"]

    # Edit the VFR data block and fill it with the testFields fields.
    vfrBlockRange = _findCellByContent(modelSheet, "VFR data block:")
    vfrBlockRange = modelSheet.cell(row=vfrBlockRange.row, column=vfrBlockRange.column+1).value
    vfrBlock      = modelSheet[vfrBlockRange]
    rowStart      = vfrBlock[0][0].row
    _substituteExcelVariable(destinySheet, rowStart, rowStart+len(vfrBlock), {"testFields": testFields})

    # Fetch the "test" header and the "iteration" block from the model.
    testBlockRange = _findCellByContent(modelSheet, "Test block:")
    testBlockRange = modelSheet.cell(row=testBlockRange.row, column=testBlockRange.column+1).value
    iterationBlockRange = _findCellByContent(modelSheet, "Iteration block:")
    iterationBlockRange = modelSheet.cell(row=iterationBlockRange.row, column=iterationBlockRange.column+1).value
    
    testBlock       = modelSheet[testBlockRange]
    iterationBlock  = modelSheet[iterationBlockRange]

    exportItems = [it for it in items if it.enabled]
    totalTestCount = len(exportItems)

    rowStart = _findCellByContent(modelSheet, "COMMENCE TESTS").row + 1

    for itemNumber, item in enumerate(exportItems):
        # Copy the test block.
        _copyRangeToStartingCell(destinySheet, testBlock, modelSheet.cell(row=rowStart, column=1))

        # Edit the newly pasted fields.
        envVars = {
            "TestResult"       : TestResult,
            "totalTestCount"   : totalTestCount,
            "itemNumber"       : itemNumber,
            "testFields"       : testFields,
            "item"             : item,
        }
        _substituteExcelVariable(destinySheet, rowStart, rowStart+len(testBlock), envVars)

        # Modify the new row with the number of pasted rows.
        rowStart += len(testBlock)

        for iterationNumber, iteration in enumerate(item.testOutput):
            # Copy the iteration block.
            _copyRangeToStartingCell(destinySheet, iterationBlock, destinySheet.cell(row=rowStart, column=1))

            # Edit the newly pasted fields. 
            newEnvVars = {
                "iterationNumber"  : iterationNumber,
                "iteration"        : iteration,
            }
            _substituteExcelVariable(destinySheet, rowStart, rowStart+len(iterationBlock), envVars | newEnvVars)

            # Modify the new row with the number of pasted rows.
            rowStart += len(iterationBlock)
    
    # Clear the "Delete Area" (where the cell range indicators are).
    deleteAreaRange = _findCellByContent(destinySheet, "Delete area:")
    deleteAreaRange = destinySheet.cell(row=deleteAreaRange.row, column=deleteAreaRange.column+1).value
    deleteBlock     = destinySheet[deleteAreaRange]
    _deleteCellRange(deleteBlock)

    # Save the modified destiny workbook.
    destinyWorkbook.save(filePath)

    # Change the print area.
    _updatePrintArea(filePath)

def _deleteCellRange(cells):
    for row in cells:
        for cell in row:
            cell.value = ""

def _findCellByContent(excel: openpyxl.worksheet, searchItem: str) -> openpyxl.cell:
    for row in excel.iter_rows():
        for cell in row:
            if cell.value == searchItem:
                return cell
    
    return None

def _copyRangeToStartingCell(excel: openpyxl.worksheet, data, startingCell: openpyxl.cell):
    startingCellRow = startingCell.row
    startingCellCol = startingCell.column

    # Loop through the copy range and paste it to the new location
    for i, row in enumerate(data):
        for j in range(len(row)):
            cell = data[i][j]
            targetCell = excel.cell(row=startingCellRow + i, column=startingCellCol + j)

            # Set value
            targetCell.value = cell.value

            # Copy the font
            if cell.font:
                targetCell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    vertAlign=cell.font.vertAlign,
                    underline=cell.font.underline,
                    strike=cell.font.strike,
                    color=cell.font.color
                )
            
            # Copy the fill (background color)
            if cell.fill:
                targetCell.fill = PatternFill(
                    fill_type=cell.fill.fill_type,
                    start_color=cell.fill.start_color,
                    end_color=cell.fill.end_color
                )
            
            # Copy the border
            if cell.border:
                targetCell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
            
            # Copy the alignment
            if cell.alignment:
                targetCell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    text_rotation=cell.alignment.text_rotation,
                    wrap_text=cell.alignment.wrap_text,
                    shrink_to_fit=cell.alignment.shrink_to_fit,
                    indent=cell.alignment.indent
                )
            
            # Copy the protection
            if cell.protection:
                targetCell.protection = Protection(
                    locked=cell.protection.locked,
                    hidden=cell.protection.hidden
                )

# This function substitutes the values inputted on the excel file by the real Python variables.
def _substituteExcelVariable(excel: openpyxl.worksheet, rowStart: int, rowEnd: int, args):
    for row in excel.iter_rows(min_row=rowStart, max_row = rowEnd-1):
        for cell in row:
            value = cell.value
            if value is None:
                continue

            matches = re.findall(r'<(.*?)>', value)
            for snippet in matches:
                value = re.sub(r'<.*?>', _evaluateVariable(snippet, args), value, count=1)

            cell.value = value

def _evaluateVariable(code: str, args):
    # Removes the possibility to import libraries or use external functions.
    args["__builtins__"] = {}
    # Variable to store the output of the code.
    args["out"] = ""
    code = "out = " + code
    
    # Execute the code using the environment given by args.
    exec(code, args)
    out = str(args["out"])

    # Remove all <> from the output because that will confound the substituteExcelVariable function.
    return re.sub(r'<.*?>', "", out)

def _updatePrintArea(xlsxPath, newArea="VFR!$A:$B"):
    # Create a temporary directory to extract files.
    tempDir = "temp_xlsx"
    if os.path.exists(tempDir):
        shutil.rmtree(tempDir)
    os.makedirs(tempDir)

    # Extract the .xlsx file contents to the temporary directory.
    with zipfile.ZipFile(xlsxPath, 'r') as zipFile:
        zipFile.extractall(tempDir)

    # Path to the workbook.xml file. This file contains the print area.
    wbXmlPath = os.path.join(tempDir, 'xl', 'workbook.xml')

    # Parse and modify the workbook.xml file.
    tree = ET.parse(wbXmlPath)
    root = tree.getroot()
    namespaces = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

    # Find the definedName element with name="_xlnm.Print_Area".
    fieldFound = False
    for definedNameField in root.findall('.//ns:definedName', namespaces):
        if definedNameField.get('name') == '_xlnm.Print_Area':
            # Update its text to the new value.
            definedNameField.text = newArea
            fieldFound = True
            break

    if not fieldFound:
        print("Set the print area on the excel model.")

    # Write the modified XML back to workbook.xml.
    tree.write(wbXmlPath, encoding='utf-8', xml_declaration=True)

    # Save file with the modified content.
    with zipfile.ZipFile(xlsxPath, 'w') as new_zip:
        for foldername, _, filenames in os.walk(tempDir):
            for filename in filenames:
                file_path = os.path.join(foldername, filename)
                arcname = os.path.relpath(file_path, tempDir)
                new_zip.write(file_path, arcname)

    # Cleanup the temporary directory.
    shutil.rmtree(tempDir)
